"""Excel workbook context and dataset scanning.

The Excel client is responsible for telling the backend which worksheet and cell
is active. This module then performs deterministic workbook inspection with
openpyxl. No LLM is involved in finding the dataset boundary.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any
import io
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter


_CELL_RE = re.compile(r"^([A-Za-z]{1,3})([1-9][0-9]*)$")


class ExcelContextError(ValueError):
    pass


@dataclass
class DatasetRegion:
    sheet_name: str
    address: str
    min_row: int
    min_col: int
    max_row: int
    max_col: int
    header_row: int
    active_cell: str | None
    columns: list[str]
    row_count: int

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheet_name": self.sheet_name,
            "address": self.address,
            "min_row": self.min_row,
            "min_col": self.min_col,
            "max_row": self.max_row,
            "max_col": self.max_col,
            "header_row": self.header_row,
            "active_cell": self.active_cell,
            "columns": self.columns,
            "row_count": self.row_count,
        }


def _is_value(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or value.strip() != "")


def _normalise_sheet_name(workbook, sheet_name: str | None) -> str:
    names = workbook.sheetnames
    if not names:
        raise ExcelContextError("Workbook contains no worksheets.")
    if not sheet_name:
        return workbook.active.title
    if sheet_name in names:
        return sheet_name
    lowered = {name.casefold(): name for name in names}
    if sheet_name.casefold() in lowered:
        return lowered[sheet_name.casefold()]
    raise ExcelContextError(
        f"Worksheet '{sheet_name}' was not found. Available sheets: {names}"
    )


def _parse_cell(address: str | None) -> tuple[int, int] | None:
    if not address:
        return None
    address = address.replace("$", "").strip()
    match = _CELL_RE.match(address)
    if not match:
        raise ExcelContextError(f"Invalid active cell '{address}'. Expected a cell such as D147.")
    from openpyxl.utils.cell import column_index_from_string
    return int(match.group(2)), column_index_from_string(match.group(1))


def _nonempty_bounds(ws) -> tuple[int, int, int, int] | None:
    min_row = min_col = None
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if _is_value(cell.value):
                r, c = cell.row, cell.column
                min_row = r if min_row is None else min(min_row, r)
                min_col = c if min_col is None else min(min_col, c)
                max_row = max(max_row, r)
                max_col = max(max_col, c)
    if min_row is None:
        return None
    return min_row, min_col, max_row, max_col


def _row_has_value(ws, row: int, min_col: int, max_col: int) -> bool:
    return any(_is_value(ws.cell(row, col).value) for col in range(min_col, max_col + 1))


def _col_has_value(ws, col: int, min_row: int, max_row: int) -> bool:
    return any(_is_value(ws.cell(row, col).value) for row in range(min_row, max_row + 1))


def _current_region(ws, row: int, col: int, bounds: tuple[int, int, int, int]) -> tuple[int, int, int, int]:
    """Find the Excel-like contiguous region around a selected cell.

    A blank row/column is treated as a dataset boundary. If the active cell is
    blank, the nearest non-empty cell is used as an anchor before expanding.
    """
    bmin_r, bmin_c, bmax_r, bmax_c = bounds
    row = max(bmin_r, min(row, bmax_r))
    col = max(bmin_c, min(col, bmax_c))

    if not _is_value(ws.cell(row, col).value):
        best = None
        best_distance = None
        for r in range(bmin_r, bmax_r + 1):
            for c in range(bmin_c, bmax_c + 1):
                if _is_value(ws.cell(r, c).value):
                    distance = abs(r - row) + abs(c - col)
                    if best_distance is None or distance < best_distance:
                        best, best_distance = (r, c), distance
        if best:
            row, col = best

    min_r = max_r = row
    min_c = max_c = col

    while min_r > bmin_r and _row_has_value(ws, min_r - 1, min_c, max_c):
        min_r -= 1
    while max_r < bmax_r and _row_has_value(ws, max_r + 1, min_c, max_c):
        max_r += 1
    while min_c > bmin_c and _col_has_value(ws, min_c - 1, min_r, max_r):
        min_c -= 1
    while max_c < bmax_c and _col_has_value(ws, max_c + 1, min_r, max_r):
        max_c += 1

    # Expanding columns can expose additional populated rows and vice versa.
    changed = True
    while changed:
        old = (min_r, min_c, max_r, max_c)
        while min_r > bmin_r and _row_has_value(ws, min_r - 1, min_c, max_c):
            min_r -= 1
        while max_r < bmax_r and _row_has_value(ws, max_r + 1, min_c, max_c):
            max_r += 1
        while min_c > bmin_c and _col_has_value(ws, min_c - 1, min_r, max_r):
            min_c -= 1
        while max_c < bmax_c and _col_has_value(ws, max_c + 1, min_r, max_r):
            max_c += 1
        changed = old != (min_r, min_c, max_r, max_c)

    return min_r, min_c, max_r, max_c


def _header_row(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> int:
    """Choose the most plausible header row from the first few rows.

    Prefer the first row with mostly non-empty, mostly unique text values.
    This avoids assuming row 1 is always the header when a sheet has a title.
    """
    limit = min(max_row, min_row + 9)
    best_row = min_row
    best_score = -1.0
    for r in range(min_row, limit + 1):
        values = [ws.cell(r, c).value for c in range(min_col, max_col + 1)]
        nonempty = [v for v in values if _is_value(v)]
        if not nonempty:
            continue
        text_count = sum(isinstance(v, str) for v in nonempty)
        unique = len({str(v).strip().casefold() for v in nonempty})
        fill = len(nonempty) / len(values)
        uniqueness = unique / len(nonempty)
        score = fill * 0.45 + (text_count / len(nonempty)) * 0.35 + uniqueness * 0.20
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def scan_workbook(
    raw_bytes: bytes,
    filename: str,
    sheet_name: str | None = None,
    active_cell: str | None = None,
    requested_range: str | None = None,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Scan one Excel workbook and return the dataset relevant to the context."""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ExcelContextError("Context-aware workbook scanning currently requires .xlsx or .xlsm.")

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=False, data_only=True)
    selected_sheet = _normalise_sheet_name(wb, sheet_name)
    ws = wb[selected_sheet]
    bounds = _nonempty_bounds(ws)
    if bounds is None:
        raise ExcelContextError(f"Worksheet '{selected_sheet}' is empty.")

    parsed = _parse_cell(active_cell)
    if requested_range:
        try:
            rmin_c, rmin_r, rmax_c, rmax_r = range_boundaries(requested_range.replace("$", ""))
        except ValueError as exc:
            raise ExcelContextError(f"Invalid dataset range '{requested_range}'.") from exc
        min_r, min_c, max_r, max_c = rmin_r, rmin_c, rmax_r, rmax_c
    elif parsed:
        min_r, min_c, max_r, max_c = _current_region(ws, parsed[0], parsed[1], bounds)
    else:
        min_r, min_c, max_r, max_c = bounds

    header = _header_row(ws, min_r, min_c, max_r, max_c)
    raw_headers = [ws.cell(header, c).value for c in range(min_c, max_c + 1)]
    columns: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw_headers, start=1):
        name = str(value).strip() if _is_value(value) else f"Unnamed_{index}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        columns.append(name)

    records = []
    for r in range(header + 1, max_r + 1):
        values = [ws.cell(r, c).value for c in range(min_c, max_c + 1)]
        if any(_is_value(v) for v in values):
            records.append(values)
    df = pd.DataFrame(records, columns=columns)

    address = f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"
    context = DatasetRegion(
        sheet_name=selected_sheet,
        address=address,
        min_row=min_r,
        min_col=min_c,
        max_row=max_r,
        max_col=max_c,
        header_row=header,
        active_cell=active_cell,
        columns=columns,
        row_count=len(df),
    ).as_dict()
    context["available_sheets"] = wb.sheetnames
    context["workbook_active_sheet"] = wb.active.title
    context["selection_inside_dataset"] = bool(
        parsed and min_r <= parsed[0] <= max_r and min_c <= parsed[1] <= max_c
    )
    return df, context
