import io

import pandas as pd
from openpyxl import Workbook
from fastapi.testclient import TestClient

from main import app
from secure_excel.semantic_roles import build_schema_profile
from secure_excel.service import interpret_query, load_excel_session, execute_query


def _restaurant_workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["RestaurantName", "Area", "Rating", "OnlineTableBooking", "OnlineDelivery"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(1, idx, header)
    rows = [
        ["Aroma", "Kolkata", 4.3, "Yes", "No"],
        ["Spice", "Kolkata", 4.8, "No", "Yes"],
        ["Dosa", "Delhi", 3.9, "Yes", "Yes"],
    ]
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _session():
    df = pd.DataFrame(
        {
            "RestaurantName": ["Aroma", "Spice", "Dosa"],
            "Area": ["Kolkata", "Kolkata", "Delhi"],
            "Rating": [4.3, 4.8, 3.9],
            "OnlineTableBooking": ["Yes", "No", "Yes"],
            "OnlineDelivery": ["No", "Yes", "Yes"],
        }
    )
    schema = build_schema_profile(df)
    from secure_excel.session_store import SESSION_STORE

    return SESSION_STORE.create(df, schema, {})


def test_schema_anonymizes_columns_to_c1_c2():
    df = pd.DataFrame({"RestaurantName": ["A", "B"], "Area": ["Kolkata", "Delhi"]})
    schema = build_schema_profile(df)
    assert [col["column_id"] for col in schema["columns"]] == ["c1", "c2"]
    assert schema["columns"][0]["role"] == "restaurant_entity"
    assert schema["columns"][1]["role"] == "geographic_area"


def test_interpret_and_execute_restaurant_delivery_query():
    session = _session()
    query = interpret_query(session.session_id, "show restaurants in Kolkata having online delivery")
    assert query["operation"] == "filter"
    assert any(cond["operator"] == "equals" and cond["value"] == "Kolkata" for cond in query["conditions"])
    assert any(cond["value"] is True for cond in query["conditions"])

    result = execute_query(session.session_id, "show restaurants in Kolkata having online delivery")
    assert result["result"]["row_count"] == 1


def test_interpret_restaurant_booking_and_rating():
    session = _session()
    booking = interpret_query(session.session_id, "show restaurants in Kolkata having online table booking")
    assert any(cond["column_id"] == "c4" and cond["value"] is True for cond in booking["conditions"])

    rated = interpret_query(session.session_id, "show restaurants with rating above 4")
    assert any(cond["column_id"] == "c3" and cond["operator"] == "greater_than" and cond["value"] == 4.0 for cond in rated["conditions"])

    combined = interpret_query(
        session.session_id,
        "show restaurants in Kolkata with rating above 4 and online delivery",
    )
    assert len(combined["conditions"]) >= 3


def test_fastapi_routes_exist():
    client = TestClient(app)
    assert client.get("/ping").status_code == 200
    assert client.get("/excel/ping").status_code == 200
    assert client.get("/powerbi/ping").status_code == 200
    assert client.get("/powerbi/transform/list").status_code == 200

