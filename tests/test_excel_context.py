import io

from openpyxl import Workbook

from common.excel_context import scan_workbook


def _workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Sales Report"
    headers = ["OrderID", "CustomerName", "Product", "Quantity", "TotalPrice"]
    for c, value in enumerate(headers, 1):
        ws.cell(3, c, value)
    for r in range(4, 9):
        ws.cell(r, 1, f"O{r}")
        ws.cell(r, 2, f"Customer {r}")
        ws.cell(r, 3, "Laptop")
        ws.cell(r, 4, r - 3)
        ws.cell(r, 5, (r - 3) * 100)
    other = wb.create_sheet("Customers")
    other.append(["CustomerID", "Name"])
    other.append([1, "Alice"])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_active_cell_finds_surrounding_dataset_and_header():
    df, context = scan_workbook(
        _workbook_bytes(),
        "sales.xlsx",
        sheet_name="Sales",
        active_cell="D6",
    )
    assert context["sheet_name"] == "Sales"
    assert context["active_cell"] == "D6"
    assert context["address"] == "A3:E8"
    assert context["header_row"] == 3
    assert context["available_sheets"] == ["Sales", "Customers"]
    assert list(df.columns) == ["OrderID", "CustomerName", "Product", "Quantity", "TotalPrice"]
    assert len(df) == 5
    assert context["selection_inside_dataset"] is True


def test_named_sheet_is_case_insensitive():
    _, context = scan_workbook(
        _workbook_bytes(),
        "sales.xlsx",
        sheet_name="sales",
        active_cell="B4",
    )
    assert context["sheet_name"] == "Sales"


def test_explicit_range_is_honoured():
    df, context = scan_workbook(
        _workbook_bytes(),
        "sales.xlsx",
        sheet_name="Sales",
        active_cell="D6",
        requested_range="B3:D8",
    )
    assert context["address"] == "B3:D8"
    assert list(df.columns) == ["CustomerName", "Product", "Quantity"]
    assert len(df) == 5
